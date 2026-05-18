from __future__ import annotations

import base64
import hashlib
import hmac
import json
import os
import secrets
import time
from dataclasses import dataclass
from pathlib import Path
from threading import Lock
from typing import Any

from fastapi import Cookie, FastAPI, HTTPException, Query, Response, status
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("DISP_DATA_DIR", BASE_DIR / "data")).resolve()
USERS_FILE = Path(os.getenv("DISP_USERS_FILE", BASE_DIR / "users.json")).resolve()
SESSION_SECRET = os.getenv("DISP_SESSION_SECRET", "")
SESSION_COOKIE_NAME = os.getenv("DISP_SESSION_COOKIE", "disp_session")
SESSION_TTL_SECONDS = int(os.getenv("DISP_SESSION_TTL", "28800"))
SESSION_COOKIE_SECURE = os.getenv("DISP_SESSION_SECURE", "true").lower() == "true"

VISIBLE_EXTENSIONS = {".xlsx"}
HEADER_MAP = {
    "enp": "enp",
    "fam": "fam",
    "im": "im",
    "ot": "ot",
    "dr": "dr",
    "lpu": "lpu",
    "date_z_1": "date_z_1",
    "date_z_2": "date_z_2",
    "disp": "disp",
}


class LoginPayload(BaseModel):
    username: str
    password: str


@dataclass
class CachedTable:
    path: Path
    mtime: float
    rows: list[dict[str, str]]


@dataclass
class CachedUsers:
    path: Path
    mtime: float
    users: dict[str, str]


class TableRepository:
    def __init__(self, data_dir: Path) -> None:
        self.data_dir = data_dir
        self._cache: dict[str, CachedTable] = {}
        self._lock = Lock()

    def list_tables(self) -> list[dict[str, str]]:
        if not self.data_dir.exists():
            return []

        items: list[dict[str, str]] = []
        for path in sorted(self.data_dir.iterdir()):
            if path.is_file() and path.suffix.lower() in VISIBLE_EXTENSIONS:
                items.append({"id": path.name, "name": path.stem.replace("_", " ")})
        return items

    def search_by_enp(self, table_id: str, enp: str) -> list[dict[str, str]]:
        normalized_enp = self._normalize_enp(enp)
        rows = self._load_rows(table_id)
        return [row for row in rows if self._normalize_enp(row.get("enp", "")) == normalized_enp]

    def get_default_table_id(self) -> str:
        tables = self.list_tables()
        if not tables:
            raise HTTPException(status_code=404, detail="На сервере не найдено ни одной таблицы.")
        return tables[0]["id"]

    def _load_rows(self, table_id: str) -> list[dict[str, str]]:
        path = (self.data_dir / table_id).resolve()
        if self.data_dir not in path.parents or not path.exists():
            raise HTTPException(status_code=404, detail="Таблица не найдена.")

        if path.suffix.lower() not in VISIBLE_EXTENSIONS:
            raise HTTPException(status_code=400, detail="Поддерживаются только .xlsx файлы.")

        stat = path.stat()
        with self._lock:
            cached = self._cache.get(table_id)
            if cached and cached.mtime == stat.st_mtime:
                return cached.rows

            rows = self._read_workbook(path)
            self._cache[table_id] = CachedTable(path=path, mtime=stat.st_mtime, rows=rows)
            return rows

    def _read_workbook(self, path: Path) -> list[dict[str, str]]:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            raw_rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not raw_rows:
            return []

        header_row = [self._normalize_header(value) for value in raw_rows[0]]
        indexes: dict[str, int] = {}
        for index, header in enumerate(header_row):
            mapped = HEADER_MAP.get(header)
            if mapped:
                indexes[mapped] = index

        missing = [name for name in HEADER_MAP.values() if name not in indexes]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"В файле {path.name} отсутствуют обязательные колонки: {', '.join(missing)}",
            )

        parsed_rows: list[dict[str, str]] = []
        for raw_row in raw_rows[1:]:
            row = list(raw_row)
            item = {
                "enp": self._string_value(row, indexes["enp"]),
                "fam": self._string_value(row, indexes["fam"]),
                "im": self._string_value(row, indexes["im"]),
                "ot": self._string_value(row, indexes["ot"]),
                "DR": self._date_value(row, indexes["dr"]),
                "lpu": self._string_value(row, indexes["lpu"]),
                "DATE_Z_1": self._date_value(row, indexes["date_z_1"]),
                "DATE_Z_2": self._date_value(row, indexes["date_z_2"]),
                "DISP": self._string_value(row, indexes["disp"]),
            }
            if item["enp"]:
                parsed_rows.append(item)

        return parsed_rows

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return str(value or "").strip().lower()

    @staticmethod
    def _normalize_enp(value: str) -> str:
        return "".join(ch for ch in str(value) if ch.isdigit())

    @staticmethod
    def _string_value(row: list[Any], index: int) -> str:
        if index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _date_value(row: list[Any], index: int) -> str:
        if index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%d.%m.%Y")
        return str(value).strip()


class UserRepository:
    def __init__(self, users_file: Path) -> None:
        self.users_file = users_file
        self._cache: CachedUsers | None = None
        self._lock = Lock()

    def verify_credentials(self, username: str, password: str) -> bool:
        users = self._load_users()
        stored_password = users.get(username)
        if stored_password is None:
            return False
        return secrets.compare_digest(password, stored_password)

    def has_user(self, username: str) -> bool:
        return username in self._load_users()

    def validate_startup(self) -> None:
        users = self._load_users()
        if not users:
            raise RuntimeError(f"Users file {self.users_file} does not contain any valid users.")

    def _load_users(self) -> dict[str, str]:
        if not self.users_file.exists():
            raise RuntimeError(f"Users file not found: {self.users_file}")

        stat = self.users_file.stat()
        with self._lock:
            if self._cache and self._cache.mtime == stat.st_mtime:
                return self._cache.users

            users = self._read_users_file()
            self._cache = CachedUsers(path=self.users_file, mtime=stat.st_mtime, users=users)
            return users

    def _read_users_file(self) -> dict[str, str]:
        try:
            payload = json.loads(self.users_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"Users file {self.users_file} contains invalid JSON.") from exc

        raw_users = payload.get("users")
        if not isinstance(raw_users, list):
            raise RuntimeError(f"Users file {self.users_file} must contain a 'users' array.")

        users: dict[str, str] = {}
        for item in raw_users:
            if not isinstance(item, dict):
                continue

            username = str(item.get("username", "")).strip()
            password = str(item.get("password", ""))
            if username and password:
                users[username] = password

        return users


repository = TableRepository(DATA_DIR)
user_repository = UserRepository(USERS_FILE)
app = FastAPI(title="Disp Registry API")

allowed_origins = [origin.strip() for origin in os.getenv("DISP_ALLOWED_ORIGINS", "*").split(",") if origin.strip()]
if not SESSION_SECRET:
    raise RuntimeError("Missing required auth setting: DISP_SESSION_SECRET")
user_repository.validate_startup()
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins or ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _sign_session_payload(payload: str) -> str:
    return hmac.new(SESSION_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()


def create_session_token(username: str) -> str:
    expires_at = int(time.time()) + SESSION_TTL_SECONDS
    nonce = secrets.token_urlsafe(12)
    payload = f"{username}|{expires_at}|{nonce}"
    signature = _sign_session_payload(payload)
    token = f"{payload}|{signature}"
    return base64.urlsafe_b64encode(token.encode("utf-8")).decode("ascii")


def read_session_token(token: str | None) -> str:
    if not token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Требуется вход.")

    try:
        decoded = base64.urlsafe_b64decode(token.encode("ascii")).decode("utf-8")
        username, expires_at_raw, nonce, signature = decoded.split("|", 3)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Сессия недействительна.") from exc

    payload = f"{username}|{expires_at_raw}|{nonce}"
    expected_signature = _sign_session_payload(payload)
    if not secrets.compare_digest(signature, expected_signature):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Сессия недействительна.")

    if not user_repository.has_user(username):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Сессия недействительна.")

    if int(expires_at_raw) < int(time.time()):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Сессия истекла.")

    return username


def require_session(session_token: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> str:
    return read_session_token(session_token)


def set_session_cookie(response: Response, username: str) -> None:
    response.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=create_session_token(username),
        max_age=SESSION_TTL_SECONDS,
        httponly=True,
        secure=SESSION_COOKIE_SECURE,
        samesite="lax",
        path="/",
    )


def clear_session_cookie(response: Response) -> None:
    response.delete_cookie(key=SESSION_COOKIE_NAME, path="/")


@app.post("/api/login")
def login(payload: LoginPayload, response: Response) -> dict[str, str]:
    if not user_repository.verify_credentials(payload.username, payload.password):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Неверный логин или пароль.")

    set_session_cookie(response, payload.username)
    response.headers["Cache-Control"] = "no-store"
    return {"status": "ok"}


@app.post("/api/logout")
def logout(response: Response) -> dict[str, str]:
    clear_session_cookie(response)
    response.headers["Cache-Control"] = "no-store"
    return {"status": "ok"}


@app.get("/api/session")
def session(response: Response, session_token: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> dict[str, bool]:
    response.headers["Cache-Control"] = "no-store"
    try:
        read_session_token(session_token)
        return {"authenticated": True}
    except HTTPException:
        return {"authenticated": False}


@app.get("/api/session/check", status_code=204)
def session_check(session_token: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> Response:
    require_session(session_token)
    return Response(status_code=204, headers={"Cache-Control": "no-store"})


@app.get("/api/health")
def health(response: Response, session_token: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> dict[str, str]:
    response.headers["Cache-Control"] = "no-store"
    require_session(session_token)
    return {"status": "ok"}


@app.get("/api/tables")
def tables(response: Response, session_token: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> list[dict[str, str]]:
    response.headers["Cache-Control"] = "no-store"
    require_session(session_token)
    return repository.list_tables()


@app.get("/api/search")
def search(
    response: Response,
    table: str | None = Query(default=None),
    enp: str = Query(..., min_length=1),
    session_token: str | None = Cookie(default=None, alias=SESSION_COOKIE_NAME),
) -> dict[str, list[dict[str, str]]]:
    response.headers["Cache-Control"] = "no-store"
    require_session(session_token)
    table_id = table or repository.get_default_table_id()
    items = repository.search_by_enp(table_id, enp)
    return {"items": items}
